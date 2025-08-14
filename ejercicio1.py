import numpy as np
import matplotlib.pyplot as plt
import cv2
import os
from skimage import measure
from skimage.filters import threshold_otsu
import pandas as pd
from tqdm import tqdm
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile

def procesar_imagen(ruta_imagen, escala=4.13, referencia_pixel=(144, 131)):
    """Procesa una imagen individual y extrae su contorno y centroide.

    Args:
        ruta_imagen (str): Ruta a la imagen a procesar
        escala (float): Factor de conversión de píxeles a micrómetros (µm/px)
        referencia_pixel (tuple): (x_ref, y_ref) píxel que define la línea y=0

    Returns:
        tuple: (contorno_µm, (centro_x, centro_y), ruta_imagen_procesada)
               o (None, None, None) si hay error o la gota no está correctamente
               identificada por encima de la referencia
    """
    try:
        # Validar parámetros de entrada
        if escala <= 0:
            raise ValueError("La escala debe ser un valor positivo")

        if not os.path.exists(ruta_imagen):
            raise FileNotFoundError(f"Imagen no encontrada: {ruta_imagen}")

        # 1. Cargar imagen en escala de grises
        imagen = cv2.imread(ruta_imagen, cv2.IMREAD_GRAYSCALE)
        if imagen is None:
            raise ValueError(f"No se pudo cargar la imagen {ruta_imagen}")

        # 2. Preprocesamiento - Suavizado Gaussiano para reducir ruido
        imagen_suavizada = cv2.GaussianBlur(imagen, (5, 5), 0)

        # 3. Binarización - Método de Otsu para umbral automático
        thresh = threshold_otsu(imagen_suavizada)
        imagen_binaria = imagen_suavizada > thresh

        # 4. Detección de contornos usando marching squares
        contornos = measure.find_contours(imagen_binaria, 0.5)
        if not contornos:
            # No se encontraron contornos útiles
            return None, None, None

        # Filtrar contornos descartando aquellos que estén principalmente bajo la referencia
        x_ref, y_ref = referencia_pixel
        contornos_arriba = [c for c in contornos if np.mean(c[:, 0]) < (y_ref + 5)]

        if len(contornos_arriba) == 0:
            # Si no hay contornos claramente arriba, tomar el más alto (menor y medio)
            contorno = min(contornos, key=lambda x: np.mean(x[:, 0]))
        else:
            # Preferir el contorno más grande entre los que están arriba
            contorno = max(contornos_arriba, key=lambda x: len(x))

        # 5. Filtrar puntos por encima de la línea de referencia (el reflejo está debajo)
        mask_arriba = contorno[:, 0] < y_ref
        contorno_filtrado = contorno[mask_arriba]

        # Si la porción por encima de la referencia es muy pequeña, considerar la imagen inválida
        if contorno_filtrado.shape[0] < 10:
            return None, None, None

        # 6. Re-centrar coordenadas: definir origen en (x_ref, y_ref), con y positivo hacia arriba
        # contorno_filtrado columnas: 0->y(px), 1->x(px)
        contorno_px = np.zeros_like(contorno_filtrado)
        contorno_px[:, 1] = contorno_filtrado[:, 1] - x_ref
        contorno_px[:, 0] = y_ref - contorno_filtrado[:, 0]

        # 7. Convertir a micrómetros
        contorno_µm = contorno_px * escala

        # 8. Calcular centroide en µm
        centro_x = float(np.mean(contorno_µm[:, 1]))
        centro_y = float(np.mean(contorno_µm[:, 0]))

        # 9. Visualización para diagnóstico (guardar imagen temporal)
        fig, ax = plt.subplots(figsize=(6, 6))
        ax.imshow(imagen, cmap='gray')
        ax.plot(contorno_filtrado[:, 1], contorno_filtrado[:, 0], 'r-', linewidth=2, label='Contorno')
        ax.axhline(y=y_ref, color='y', linestyle='--', linewidth=1.0, label='Referencia y=0')
        centro_px_x = centro_x / escala + x_ref
        centro_px_y = y_ref - (centro_y / escala)
        ax.scatter(centro_px_x, centro_px_y, c='blue', marker='x', s=100, label='Centroide')
        ax.axis('off')
        plt.legend(loc='upper right')
        plt.title(f'Procesamiento: {os.path.basename(ruta_imagen)}')
        plt.tight_layout()

        temp_img = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
        plt.savefig(temp_img.name, bbox_inches='tight', pad_inches=0, dpi=150)
        plt.close()

        return contorno_µm, (centro_x, centro_y), temp_img.name

    except Exception as e:
        print(f"Error procesando {ruta_imagen}: {str(e)}")
        return None, None, None


def procesar_todas_imagenes(carpeta_imagenes, num_imagenes=126, escala=4.13):
    """Procesa todas las imágenes y recolecta todos los datos con validación mejorada.

    Args:
        carpeta_imagenes (str): Ruta a la carpeta con imágenes
        num_imagenes (int): Número total de imágenes a procesar
        escala (float): Factor de conversión píxeles a micrómetros

    Returns:
        pd.DataFrame: DataFrame con todos los datos procesados
    """
    # Validación inicial
    if not os.path.isdir(carpeta_imagenes):
        raise ValueError(f"La carpeta {carpeta_imagenes} no existe")

    if num_imagenes <= 0:
        raise ValueError("El número de imágenes debe ser positivo")

    datos = []
    fps = 20538  # Frames por segundo según enunciado
    tiempos = np.arange(num_imagenes) / fps

    # Procesar con barra de progreso
    for i in tqdm(range(1, num_imagenes + 1), desc="Procesando imágenes"):
        num_str = f"{i:04d}"
        nombre_imagen = f"TP4_Gota_{num_str}.jpg"
        ruta_imagen = os.path.join(carpeta_imagenes, nombre_imagen)

        # Procesar imagen
        contorno, centro, img_path = procesar_imagen(ruta_imagen, escala)

        # Preparar datos para el DataFrame
        dato_imagen = {
            'Imagen': nombre_imagen,
            'Tiempo (s)': tiempos[i - 1],
            'Centroide_x (µm)': None,
            'Centroide_y (µm)': None,
            'N_puntos_contorno': 0,
            'Ruta_imagen_procesada': '',
            'Contorno_x': '[]',
            'Contorno_y': '[]'
        }

        if contorno is not None and centro is not None:
            dato_imagen.update({
                'Centroide_x (µm)': centro[0],
                'Centroide_y (µm)': centro[1],
                'N_puntos_contorno': len(contorno),
                'Ruta_imagen_procesada': img_path,
                'Contorno_x': json.dumps(contorno[:, 1].tolist()),
                'Contorno_y': json.dumps(contorno[:, 0].tolist())
            })

        datos.append(dato_imagen)

    return pd.DataFrame(datos)

def exportar_a_excel(df, nombre_archivo='resultados_completos.xlsx'):
    """Exporta todos los datos a un archivo Excel en una sola hoja con mejor formato.

    Args:
        df (pd.DataFrame): DataFrame con los datos a exportar
        nombre_archivo (str): Nombre del archivo Excel de salida

    Returns:
        bool: True si la exportación fue exitosa
    """
    try:
        # Validar DataFrame
        if df.empty:
            raise ValueError("El DataFrame está vacío")

        required_cols = ['Imagen', 'Tiempo (s)', 'Centroide_x (µm)', 'Centroide_y (µm)',
                         'N_puntos_contorno', 'Contorno_x', 'Contorno_y']
        if not all(col in df.columns for col in required_cols):
            raise ValueError(f"DataFrame no contiene columnas requeridas: {required_cols}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Completos"

        header_style = Font(bold=True, color="FFFFFF")
        fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Escribir encabezado
        columnas = ['Imagen', 'Tiempo (s)', 'Centroide_x (µm)', 'Centroide_y (µm)',
                    'N_puntos_contorno', 'Contorno_x', 'Contorno_y']

        for col_num, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = header_style
            cell.fill = fill

        for row_num, row_data in enumerate(dataframe_to_rows(df[columnas], index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        column_widths = {
            'A': 20,  # Imagen
            'B': 12,  # Tiempo
            'C': 16,  # Centroide X
            'D': 16,  # Centroide Y
            'E': 18,  # N_puntos
            'F': 30,  # Contorno_x
            'G': 30  # Contorno_y
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Aplicar formato numérico y alineación
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                if cell.column_letter in ['B', 'C', 'D', 'E']:
                    cell.number_format = '0.00'

        wb.save(nombre_archivo)

        print(f"\nArchivo Excel generado exitosamente: {nombre_archivo}")
        return True

    except Exception as e:
        print(f"\nError al exportar a Excel: {str(e)}")
        return False

def graficar_centroides_vs_tiempo(df, nombre_archivo='centroides_vs_tiempo.png'):
    """Grafica la posición X e Y del centroide con estilo mejorado."""
    try:
        plt.figure(figsize=(14, 6))

        # Configuración de estilo
        plt.style.use('seaborn-v0_8')
        plt.rcParams['axes.grid'] = True
        plt.rcParams['grid.alpha'] = 0.3

        # Gráfico para X
        plt.subplot(1, 2, 1)
        plt.plot(df['Tiempo (s)'], df['Centroide_x (µm)'], 'b-', linewidth=2,
                 label='Posición X')
        plt.xlabel('Tiempo (s)', fontsize=12)
        plt.ylabel('Posición X (µm)', fontsize=12)
        plt.title('Evolución de la posición X del centroide', fontsize=14, pad=20)
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.legend(fontsize=10)

        # Gráfico para Y
        plt.subplot(1, 2, 2)
        plt.plot(df['Tiempo (s)'], df['Centroide_y (µm)'], 'r-', linewidth=2,
                 label='Posición Y')
        plt.xlabel('Tiempo (s)', fontsize=12)
        plt.ylabel('Posición Y (µm)', fontsize=12)
        plt.title('Evolución de la posición Y del centroide', fontsize=14, pad=20)
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.legend(fontsize=10)

        plt.tight_layout()
        plt.savefig(nombre_archivo, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"\nGráfico de centroides guardado como: {nombre_archivo}")

    except Exception as e:
        print(f"\nError al graficar centroides: {str(e)}")


def generar_informe1(carpeta_imagenes, num_imagenes=126):
    """Función principal mejorada con manejo de errores robusto."""
    print("\n--- Ejercicio 1: Procesamiento de imágenes ---")

    try:
        # Validación inicial
        if not os.path.exists(carpeta_imagenes):
            raise FileNotFoundError(f"La carpeta '{carpeta_imagenes}' no existe")

        if num_imagenes <= 0:
            raise ValueError("El número de imágenes debe ser positivo")

        print("\nIniciando procesamiento de imágenes...")
        df = procesar_todas_imagenes(carpeta_imagenes, num_imagenes)

        if df.empty:
            raise ValueError("No se pudieron procesar imágenes (¿formato incorrecto?)")

        # Análisis de resultados
        print("\nResumen estadístico:")
        print(f"- Imágenes procesadas: {len(df)}")
        print(f"- Centroide X promedio: {df['Centroide_x (µm)'].mean():.2f} µm")
        print(f"- Centroide Y promedio: {df['Centroide_y (µm)'].mean():.2f} µm")

        # Exportar resultados
        print("\nExportando resultados a Excel...")
        if exportar_a_excel(df):
            print("\nGenerando gráficos de análisis...")
            graficar_centroides_vs_tiempo(df)

            # Gráfico adicional de trayectoria del centroide
            plt.figure(figsize=(8, 6))
            plt.plot(df['Centroide_x (µm)'], df['Centroide_y (µm)'],
                     'b-', linewidth=1.5, label='Trayectoria')
            plt.xlabel('Posición X (µm)', fontsize=12)
            plt.ylabel('Posición Y (µm)', fontsize=12)
            plt.title('Trayectoria del centroide de la gota', fontsize=14)
            plt.grid(True, linestyle='--', alpha=0.3)
            plt.legend()
            plt.tight_layout()
            plt.savefig('trayectoria_centroide.png', dpi=300)
            plt.close()

            print("\nProceso completado exitosamente!")
            return True
        else:
            print("\nProceso completado con errores.")
            return False

    except Exception as e:
        print(f"\nERROR durante el procesamiento: {str(e)}")
        return False